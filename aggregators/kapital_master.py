import os
import pandas as pd
from openpyxl import Workbook
import re

DATA_DIR = "processed_data/kapital_bank_excel"
OUTPUT_DIR = "master_data"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Find all .xlsx files in DATA_DIR and its subfolders
FILES = []
for root, dirs, files in os.walk(DATA_DIR):
    for f in files:
        if f.endswith('.xlsx'):
            FILES.append(os.path.join(root, f))

print(f"Files found: {FILES}")

master_wb = Workbook()
default_sheet = master_wb.active
sheets_added = 0

for file_path in FILES:
    file = os.path.basename(file_path)
    m = re.match(r"(.+?)_(\d{4})_(Q\d)\.xlsx", file)
    if not m:
        print(f"Skipping file with unexpected name: {file}")
        continue
    report_type, year, quarter = m.groups()
    xl = pd.ExcelFile(file_path)
    for idx, sheet_name in enumerate(xl.sheet_names, 1):
        df = xl.parse(sheet_name)
        print(f"Adding: {file}, sheet: {sheet_name}, shape: {df.shape}")
        if df.empty:
            continue
        sheet_title = f"{year}_{quarter}_{report_type}_{idx}"
        sheet_title = sheet_title[:31]
        ws = master_wb.create_sheet(title=sheet_title)
        for col_idx, col in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col)
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        sheets_added += 1

if sheets_added > 0:
    master_wb.remove(default_sheet)

output_path = os.path.join(OUTPUT_DIR, "master.xlsx")
master_wb.save(output_path)
print(f"✅ Master Excel saved to {output_path} with {sheets_added} sheets.")
