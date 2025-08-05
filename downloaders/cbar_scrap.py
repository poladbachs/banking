import requests
import hashlib
import os
import time
import re
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

# ---- DIRECTORIES ----
CBAR_DIR = os.path.join("processed_data", "CBAR")
os.makedirs(CBAR_DIR, exist_ok=True)

CBAR_LOCAL_FILE = os.path.join(CBAR_DIR, "CBAR_LATEST.xlsx")
CBAR_HASH_FILE = os.path.join(CBAR_DIR, "CBAR_LATEST.hash")

MONTHS_AZ = [
    "Yanvar", "Fevral", "Mart", "Aprel", "May", "İyun", "İyul",
    "Avqust", "Sentyabr", "Oktyabr", "Noyabr", "Dekabr"
]
MONTHS_EN = [
    "January", "February", "March", "April", "May", "June", "July",
    "August", "September", "October", "November", "December"
]

from openpyxl import load_workbook

CBAR_SHEETS_KEEP = [
    "2.7", "2.7.d.", "2.8", "2.8.1", "2.8.2", "2.8.3",
    "5.2", "5.3", "5.4", "5.6", "5.7", "6.1", "6.2"
]



def filter_cbar_sheets(filepath, sheets_to_keep=CBAR_SHEETS_KEEP):
    wb = load_workbook(filepath)
    sheets_all = wb.sheetnames
    for name in sheets_all:
        if name not in sheets_to_keep:
            std = wb[name]
            wb.remove(std)
    wb.save(filepath)
    wb.close()
    print(f"[INFO] Filtered to sheets: {sheets_to_keep}")


MONTH_MAP = {az: en for az, en in zip(MONTHS_AZ, MONTHS_EN)}

def download_latest_cbar_excel(download_to=CBAR_LOCAL_FILE):
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    url = "https://www.cbar.az/page-40/statistical-bulletin"
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.get(url)
    time.sleep(2)

    # Find active year
    years = driver.find_elements(By.CSS_SELECTOR, "dt")
    active_year = None
    for y in years:
        if 'jquery-list-active' in y.get_attribute("class"):
            active_year = y.text.strip()
            break
    if not active_year:
        active_year = years[0].text.strip() if years else "UnknownYear"

    # Find the latest Excel link and parse period (month)
    assets = driver.find_elements(By.CSS_SELECTOR, "dd.assets a.download_item")
    excel_url = None
    period = None
    for a in assets:
        href = a.get_attribute("href")
        if href and href.endswith(".xlsx"):
            excel_url = href
            try:
                div = a.find_element(By.TAG_NAME, "div")
                div_html = div.get_attribute('outerHTML')
                print(f"[DEBUG] Raw <div> HTML: {div_html!r}")
                m = re.search(r"<p[^>]*>(.*?)</p>", div_html, re.IGNORECASE)
                if m:
                    period = m.group(1).strip()
                    print(f"[DEBUG] Extracted month from HTML: {period!r}")
                else:
                    period = "Unknown"
            except Exception as ex:
                print(f"[DEBUG] Period extraction error: {ex}")
                period = "Unknown"
            break
    driver.quit()
    if not excel_url:
        raise Exception("Excel link not found")
    period_en = MONTH_MAP.get(period, period)
    period_full = f"{period} {active_year}"
    period_full_en = f"{period_en} {active_year}"
    print(f"[INFO] Latest period on website: {period_full} / {period_full_en}")
    print(f"[INFO] Latest Excel URL: {excel_url}")
    resp = requests.get(excel_url)
    with open(download_to + ".new", "wb") as f:
        f.write(resp.content)
    print(f"[INFO] Downloaded as: {download_to + '.new'}")
    return download_to + ".new", period_full, period_full_en

def file_hash(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()

def load_stored_hash():
    if os.path.exists(CBAR_HASH_FILE):
        with open(CBAR_HASH_FILE, "r") as f:
            return f.read().strip()
    return None

def store_hash(hashval):
    with open(CBAR_HASH_FILE, "w") as f:
        f.write(hashval)

def update_cbar_file():
    TMP_FILE, period_full, period_full_en = download_latest_cbar_excel()
    new_hash = file_hash(TMP_FILE)

    prev_hash = load_stored_hash()
    # filename for new period
    clean_period = period_full_en.replace(" ", "_")
    new_fname = f"CBAR_{clean_period}.xlsx"
    new_fpath = os.path.join(CBAR_DIR, new_fname)

    if prev_hash == new_hash and os.path.exists(new_fpath):
        print(f"[INFO] No update. CBAR Excel unchanged for period: {period_full} / {period_full_en}")
        os.remove(TMP_FILE)
    else:
        print(f"[INFO] New data detected or no previous file. Updating local CBAR file for period: {period_full} / {period_full_en}")
        # Remove all old CBAR files, only keep the latest
        for f in os.listdir(CBAR_DIR):
            if f.endswith(".xlsx"):
                os.remove(os.path.join(CBAR_DIR, f))
        os.rename(TMP_FILE, new_fpath)
        filter_cbar_sheets(new_fpath)
        store_hash(new_hash)
        print(f"[INFO] {new_fname} now up to date. Period: {period_full} / {period_full_en}")
    print(f"[INFO] Current local hash: {new_hash}")

if __name__ == "__main__":
    update_cbar_file()